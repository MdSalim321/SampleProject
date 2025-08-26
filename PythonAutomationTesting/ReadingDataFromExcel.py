import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Load Excel file
wb = openpyxl.load_workbook("DataSet/Data.xlsx")
sheet = wb.active
rows = sheet.max_row

# Launch browser
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
driver.implicitly_wait(20)
wait = WebDriverWait(driver, 20)

# Loop through all rows (skip header row 1)
for rw in range(2, rows + 1):
    userName = sheet.cell(row=rw, column=1).value
    passWord = sheet.cell(row=rw, column=2).value
    print("UserName:", userName, "| Password:", passWord)

    # Wait for fields
    user_field = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Username']")))
    pass_field = wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@placeholder='Password']")))

    # Clear & enter values
    user_field.clear()
    pass_field.clear()
    user_field.send_keys(userName)
    pass_field.send_keys(passWord)

    # Click login
    loginBtn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@type='submit']")))
    loginBtn.click()
    time.sleep(3)

    try:
        # Wait for user dropdown (login success check)
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "oxd-userdropdown-tab")))
        print("✅ Login Success for:", userName)

        # Logout
        driver.find_element(By.CLASS_NAME, "oxd-userdropdown-tab").click()
        time.sleep(2)
        driver.find_element(By.LINK_TEXT, "Logout").click()
        time.sleep(2)

    except:
        print("❌ Login Failed for:", userName)
        driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

driver.quit()
