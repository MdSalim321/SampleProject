from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
import os
import time

# Setup WebDriver
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Chrome(options=options)

try:
    # ------------------- Step 1: Login Page -------------------
    driver.get("https://practicetestautomation.com/practice-test-login/")

    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, "username")))

    driver.find_element(By.ID, "username").send_keys("student")
    driver.find_element(By.ID, "password").send_keys("Password123")
    driver.find_element(By.ID, "submit").click()

    # ------------------- Step 2: Wait for Dashboard -------------------
    success_msg = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, "post-title"))
    )
    print("Login Successful:", success_msg.text)

    # ------------------- Step 3: Upload File -------------------
    driver.get("https://the-internet.herokuapp.com/upload")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "file-upload")))

    # Create sample file
    file_path = os.path.abspath("sample.txt")
    with open(file_path, "w") as f:
        f.write("Sample file upload")

    driver.find_element(By.ID, "file-upload").send_keys(file_path)
    driver.find_element(By.ID, "file-submit").click()

    print("File uploaded successfully.")

    # ------------------- Step 4: Screenshot -------------------
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.TAG_NAME, "h3")))
    driver.save_screenshot("screenshot.png")
    print("Screenshot saved.")

    # ------------------- Step 5: Excel Read and Form Fill -------------------
    # Create sample Excel file
    excel_path = "form_data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Email"])
    ws.append(["Alice Johnson", "alice@example.com"])
    wb.save(excel_path)

    # Read from Excel
    wb = load_workbook(excel_path)
    ws = wb.active
    name = ws["A2"].value
    email = ws["B2"].value

    # Fill a form using Excel data
    driver.get("https://www.w3schools.com/html/html_forms.asp")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "fname")))

    driver.find_element(By.ID, "fname").clear()
    driver.find_element(By.ID, "fname").send_keys(name)

    driver.find_element(By.ID, "lname").clear()
    driver.find_element(By.ID, "lname").send_keys(email)

    print(f"Filled form with: Name = {name}, Email = {email}")

finally:
    time.sleep(3)
    driver.quit()
